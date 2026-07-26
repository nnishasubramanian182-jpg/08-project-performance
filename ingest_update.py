"""
Ongoing ingest script for Master Userlist + Daily Records.

Usage:
  python3 ingest_update.py --userlist lotteryUserInfo_A.xlsx lotteryUserInfo_B.xlsx
  python3 ingest_update.py --deposits water_new.xlsx
  python3 ingest_update.py --withdrawals withdraw_new.xlsx
  python3 ingest_update.py --wallet detail_new1.xlsx detail_new2.xlsx
  python3 ingest_update.py --deposits a.xlsx --withdrawals b.xlsx --wallet c.xlsx --userlist d.xlsx

Any combination of the four flags can be passed in one run. After ingest, Daily
Records tables are purged to a rolling 33-day window (by create_time), and both
DBs are re-uploaded to R2 automatically unless --no-upload is passed.
"""
import argparse
import os
import re
import sqlite3
import subprocess
import sys

BASE = os.path.dirname(os.path.abspath(__file__))
MASTER_DB = os.path.join(BASE, "master_userlist.db")
DAILY_DB = os.path.join(BASE, "daily_records.db")
RETENTION_DAYS = 33

import openpyxl


def clean(row):
    return tuple(str(v) if hasattr(v, "isoformat") else v for v in row)


def load_sheet(path):
    """Reads EVERY sheet in the workbook whose header row matches the
    first sheet's header exactly, concatenating their rows -- confirmed
    that at least the wallet detail export can split transaction rows
    across multiple sheets/pages instead of one, and reading only
    wb.active (the previous behavior) silently dropped every sheet after
    the first. A sheet whose header doesn't match (e.g. a summary/pivot
    tab) is skipped with a warning rather than blindly appended, since its
    columns wouldn't line up with the data rows anyway."""
    wb = openpyxl.load_workbook(path, read_only=True)
    sheet_names = wb.sheetnames
    first_ws = wb[sheet_names[0]]
    header = next(first_ws.iter_rows(min_row=1, max_row=1, values_only=True))
    rows = list(first_ws.iter_rows(min_row=2, values_only=True))
    for name in sheet_names[1:]:
        ws = wb[name]
        sheet_header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if sheet_header != header:
            print(f"  WARNING: {path}: sheet '{name}' has a different header, skipping (not transaction rows?)")
            continue
        extra_rows = list(ws.iter_rows(min_row=2, values_only=True))
        if extra_rows:
            print(f"  {path}: sheet '{name}' contributed {len(extra_rows)} additional rows")
        rows.extend(extra_rows)
    wb.close()
    return header, rows


def already_ingested(conn, filename):
    conn.execute("CREATE TABLE IF NOT EXISTS ingested_files (filename TEXT PRIMARY KEY, ingested_at TEXT DEFAULT CURRENT_TIMESTAMP)")
    row = conn.execute("SELECT 1 FROM ingested_files WHERE filename = ?", (os.path.basename(filename),)).fetchone()
    return row is not None


def mark_ingested(conn, filename):
    conn.execute("INSERT OR IGNORE INTO ingested_files (filename) VALUES (?)", (os.path.basename(filename),))


def ingest_agents(files):
    """Agent-to-user assignment sheet (e.g. "Agent-users.xlsx"). Unlike the
    other ingest_* functions, the source layout isn't one-row-per-user --
    each column is an agent name, and every non-blank cell below it is a
    user_id assigned to that agent. Only the "Mastersheet 04" tab is treated
    as authoritative (confirmed with the user: other tabs in the same
    workbook, like "04 (OLD)" or "Sales Team - Mastersheet from 1", are
    stale/different-team snapshots with heavily overlapping user_ids and
    conflicting agent names, not something to merge in automatically).

    If a user_id appears in more than one column of the same sheet, the
    left-most column wins (deterministic, and matches how the one such
    conflict found during initial import was resolved)."""
    conn = sqlite3.connect(MASTER_DB)
    cur = conn.cursor()
    cur.execute("CREATE TABLE IF NOT EXISTS agent_assignments (user_id INTEGER PRIMARY KEY, agent_name TEXT)")
    total_pairs, total_conflicts = 0, 0
    for f in files:
        if already_ingested(conn, f):
            print(f"  skip (already ingested): {f}")
            continue
        wb = openpyxl.load_workbook(f, read_only=True)
        sheet_name = "Mastersheet 04" if "Mastersheet 04" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        header = next(rows)
        col_agents = {i: str(h).strip() for i, h in enumerate(header) if h}
        mapping = {}
        for row in rows:
            for i, agent in col_agents.items():
                v = row[i] if i < len(row) else None
                if v is None:
                    continue
                try:
                    uid = int(float(v))
                except (TypeError, ValueError):
                    continue
                if uid in mapping and mapping[uid] != agent:
                    total_conflicts += 1
                    continue  # left-most column already claimed this user_id
                mapping.setdefault(uid, agent)
        wb.close()
        cur.executemany(
            "INSERT OR REPLACE INTO agent_assignments (user_id, agent_name) VALUES (?, ?)",
            list(mapping.items()),
        )
        total_pairs += len(mapping)
        mark_ingested(conn, f)
        conn.commit()
        print(f"  {f}: {len(mapping)} user->agent assignments from sheet '{sheet_name}' ({total_conflicts} same-sheet conflicts resolved left-most-wins)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_agent_name ON agent_assignments(agent_name)")
    conn.commit()
    print(f"Agent assignments: {total_pairs} total (user, agent) pairs processed")
    conn.close()


def ingest_bulk_reassign(files):
    """Bulk agent reassignment: a simple two-column sheet (Column A = User
    ID, Column B = Agent Name), for correcting/reassigning a specific batch
    of users at once -- unlike ingest_agents()'s wide one-column-per-agent
    "Mastersheet 04" layout, which is a full agent-list refresh.

    Every agent name in the file is validated against the names ALREADY in
    agent_assignments (the exact same list the dashboard's Reassign Agent
    dropdown is built from) BEFORE anything is written. A single typo'd
    agent name fails the WHOLE file rather than silently creating a new,
    slightly-different agent bucket that would never show up correctly
    anywhere else on the dashboard. "Un-Assigned" (case-insensitive) is
    always accepted and clears the assignment instead of setting one."""
    conn = sqlite3.connect(MASTER_DB)
    cur = conn.cursor()
    cur.execute("CREATE TABLE IF NOT EXISTS agent_assignments (user_id INTEGER PRIMARY KEY, agent_name TEXT)")
    known_agents = {row[0] for row in cur.execute("SELECT DISTINCT agent_name FROM agent_assignments").fetchall()}

    for f in files:
        if already_ingested(conn, f):
            print(f"  skip (already ingested): {f}")
            continue
        _, rows = load_sheet(f)

        parsed = []  # (row_num, user_id, agent_name-or-None-for-unassign)
        invalid = []  # (row_num, user_id, bad_agent_name)
        for i, row in enumerate(rows, start=2):
            user_id_raw = row[0] if len(row) > 0 else None
            agent_raw = row[1] if len(row) > 1 else None
            if user_id_raw is None:
                continue
            try:
                user_id = int(float(user_id_raw))
            except (TypeError, ValueError):
                invalid.append((i, user_id_raw, f"invalid User ID: {user_id_raw!r}"))
                continue
            agent_name = str(agent_raw).strip() if agent_raw else ""
            if agent_name.lower() == "un-assigned":
                parsed.append((i, user_id, None))
            elif agent_name in known_agents:
                parsed.append((i, user_id, agent_name))
            else:
                invalid.append((i, user_id, agent_name))

        if invalid:
            print(f"FATAL: {len(invalid)} row(s) in {f} have an agent name that doesn't match the dashboard:", file=sys.stderr)
            for row_num, user_id, bad_name in invalid[:50]:
                print(f"  row {row_num}: user_id={user_id} agent={bad_name!r}", file=sys.stderr)
            if len(invalid) > 50:
                print(f"  ... and {len(invalid) - 50} more", file=sys.stderr)
            print("Valid agent names (must match exactly, including WFH/SL suffix and spacing):", file=sys.stderr)
            for name in sorted(known_agents):
                print(f"  - {name}", file=sys.stderr)
            print("  - Un-Assigned", file=sys.stderr)
            conn.close()
            sys.exit(1)

        missing_users = []
        applied = 0
        for row_num, user_id, agent_name in parsed:
            exists = cur.execute("SELECT 1 FROM users WHERE user_id = ?", (user_id,)).fetchone()
            if not exists:
                missing_users.append(user_id)
                continue
            if agent_name:
                cur.execute(
                    "INSERT OR REPLACE INTO agent_assignments (user_id, agent_name) VALUES (?, ?)",
                    (user_id, agent_name),
                )
            else:
                cur.execute("DELETE FROM agent_assignments WHERE user_id = ?", (user_id,))
            applied += 1
        mark_ingested(conn, f)
        conn.commit()

        if missing_users:
            shown = missing_users[:20]
            more = f" (+{len(missing_users) - 20} more)" if len(missing_users) > 20 else ""
            print(f"  Warning: {len(missing_users)} user_id(s) not found in users table, skipped: {shown}{more}")
        print(f"  {f}: {applied} agent reassignments applied")

    conn.close()


# Columns WE maintain ourselves via the ongoing hourly sync (never present
# in the source lotteryUserInfo file, added later via ALTER TABLE). A
# userlist re-upload must NEVER reset these -- doing so wipes the
# high-water mark that stops sync_master_userlist() from re-adding
# deposits/withdrawals already reflected in the file's own total_recharge/
# total_withdrawal. That is exactly how a one-time historical double-count
# got baked into every user's lifetime totals the first time these columns
# were introduced (confirmed for user 1761219: total_recharge was inflated
# by precisely their full currently-retained deposit sum) -- and using a
# positional INSERT OR REPLACE here (the previous approach) would silently
# repeat the exact same bug on every future re-upload, since REPLACE resets
# any column not present in the VALUES list to its default (NULL).
OWN_TRACKING_COLUMNS = ("deposit_sync_time", "withdrawal_sync_time")


def ingest_userlist(files):
    conn = sqlite3.connect(MASTER_DB)
    cur = conn.cursor()
    all_cols = [r[1] for r in cur.execute("PRAGMA table_info(users)").fetchall()]
    file_cols = [c for c in all_cols if c not in OWN_TRACKING_COLUMNS]
    n_cols = len(file_cols)
    update_time_idx = n_cols - 2
    insert_cols_sql = ", ".join(file_cols)
    update_cols_sql = ", ".join(f"{c} = ?" for c in file_cols[1:])  # skip user_id (WHERE key, not SET)
    updated, inserted, skipped_files, skipped_rows = 0, 0, 0, 0
    for f in files:
        if already_ingested(conn, f):
            print(f"  skip (already ingested): {f}")
            skipped_files += 1
            continue
        _, rows = load_sheet(f)
        for row in rows:
            if row[0] is None:
                continue
            row = list(clean(row))
            if len(row) != n_cols:
                print(
                    f"  WARNING: row for user {row[0]!r} has {len(row)} columns, expected {n_cols} "
                    f"(source file layout changed?) -- skipping this row",
                    file=sys.stderr,
                )
                skipped_rows += 1
                continue
            row[0] = int(float(row[0]))
            uid = row[0]
            existing = cur.execute("SELECT update_time FROM users WHERE user_id = ?", (uid,)).fetchone()
            if existing is None:
                cur.execute(f"INSERT INTO users ({insert_cols_sql}) VALUES ({','.join(['?']*n_cols)})", row)
                inserted += 1
            else:
                new_ut, old_ut = row[update_time_idx], existing[0]
                if new_ut is not None and (old_ut is None or str(new_ut) > str(old_ut)):
                    cur.execute(f"UPDATE users SET {update_cols_sql} WHERE user_id = ?", row[1:] + [uid])
                    updated += 1
        mark_ingested(conn, f)
        conn.commit()
    print(f"Master Userlist: {inserted} new, {updated} updated, {skipped_rows} rows skipped (bad shape), {skipped_files} files already ingested")
    conn.close()


def ingest_deposits(files):
    # INSERT OR REPLACE (not IGNORE): a re-fetched deposit with the same id but a
    # changed status (e.g. pending -> COMPLETE some hours/days later) must overwrite
    # the existing row, not be silently skipped.
    conn = sqlite3.connect(DAILY_DB)
    cur = conn.cursor()
    n_cols = len(cur.execute("PRAGMA table_info(deposits)").fetchall())
    added = 0
    for f in files:
        if already_ingested(conn, f):
            print(f"  skip (already ingested): {f}")
            continue
        _, rows = load_sheet(f)
        cur.executemany(f"INSERT OR REPLACE INTO deposits VALUES ({','.join(['?']*n_cols)})", [clean(r) for r in rows])
        added += cur.rowcount
        mark_ingested(conn, f)
        conn.commit()
    print(f"Deposits: {added} rows processed (new + updated)")
    conn.close()


def ingest_withdrawals(files):
    # INSERT OR REPLACE (not IGNORE): a re-fetched withdrawal with the same id but a
    # changed status (In-Review/Processing -> Complete/Rejected/Failed, possibly days
    # later) must overwrite the existing row, not be silently skipped.
    conn = sqlite3.connect(DAILY_DB)
    cur = conn.cursor()
    n_cols = len(cur.execute("PRAGMA table_info(withdrawals)").fetchall())
    added = 0
    for f in files:
        if already_ingested(conn, f):
            print(f"  skip (already ingested): {f}")
            continue
        _, rows = load_sheet(f)
        cur.executemany(f"INSERT OR REPLACE INTO withdrawals VALUES ({','.join(['?']*n_cols)})", [clean(r) for r in rows])
        added += cur.rowcount
        mark_ingested(conn, f)
        conn.commit()
    print(f"Withdrawals: {added} rows processed (new + updated)")
    conn.close()


def normalize(s):
    return re.sub(r"\s+", " ", str(s).strip().lower())


def classify_bonus(game_name, source, source_id):
    """A wallet_transactions row is a bonus credit under any of four rules,
    all confirmed against real data:

    1. game_name is a real bonus name (e.g. "Welcome Back Bonus", "VIP
       Level: 3") AND source is BLANK -- every real game has a populated
       source (its provider, e.g. Evolution/JDB/KoolBet), every actual bonus
       category has 100% blank source. Category = game_name itself, so any
       NEW bonus name is picked up automatically with no maintenance.

    2. game_name is literally "Elle Import Excel Add" -- a generic wrapper
       label used for a second bonus family. Checked BEFORE rule 1 (which
       would otherwise match it too, but lump every row under the
       meaningless label "Elle Import Excel Add"): the real bonus identity
       lives in source_id instead (confirmed values: "Daily Active Low",
       "Daily Active Low VIP"), always with a blank source too.

    3. game_name is BLANK and source_id contains the word "bonus" -- a third
       family ("Daily Active Bonus-<random hex>", "Daily Active Bonus
       Low-<random hex>") confirmed distinct from the other blank-game_name
       rows, which carry deposit/withdrawal order-number references in
       source_id instead (e.g. "DI2026070101110003"), not bonus text. The
       per-instance random suffix is stripped so every instance rolls up
       into one combined category each, rather than ~900 near-duplicate
       ones (confirmed: 911 total split exactly 617 "Daily Active Bonus" +
       294 "Daily Active Bonus Low", no overlap)."""
    game_name = str(game_name).strip() if game_name else ""
    source = str(source).strip() if source else ""
    source_id = str(source_id).strip() if source_id else ""

    if game_name == "Elle Import Excel Add":
        return source_id or game_name

    if game_name and not source:
        return game_name

    if not game_name and "bonus" in source_id.lower():
        lowered = source_id.lower()
        if lowered.startswith("daily active bonus low"):
            return "Daily Active Bonus Low"
        if lowered.startswith("daily active bonus"):
            return "Daily Active Bonus"
        return source_id

    # 4. game_name is BLANK and source_id starts with "WEEKLY_SIGN" -- a
    # fourth family (weekly sign-in/check-in credit), rolled up into one
    # combined category the same way Daily Active Bonus is, rather than
    # exposing the raw source_id per instance.
    if not game_name and source_id.upper().startswith("WEEKLY_SIGN"):
        return "Weekly Check-IN Bonus"

    return None


# Bump this whenever classify_bonus() gains or changes a rule -- it forces
# ingest_wallet()'s backfill step to do one full re-scan of wallet_transactions
# under the new rules, then fall back to only scanning genuinely new rows.
# Without this, a rule change would only apply to rows inserted AFTER the
# change; existing rows that now match would silently stay unclassified.
CLASSIFY_BONUS_RULES_VERSION = 4


def ingest_wallet(files):
    conn = sqlite3.connect(DAILY_DB)
    cur = conn.cursor()
    n_cols = len(cur.execute("PRAGMA table_info(wallet_transactions)").fetchall())
    # bonuses is normally created once by the original bootstrap (build_daily_records.py),
    # not by this ongoing script -- IF NOT EXISTS here so a from-scratch daily_records.db
    # doesn't fail on the INSERT below.
    cur.execute(
        "CREATE TABLE IF NOT EXISTS bonuses ("
        "id INTEGER PRIMARY KEY, user_id INTEGER, bonus_name TEXT, matched_category TEXT, "
        "change_value REAL, change_after REAL, create_time TEXT, source TEXT)"
    )
    # Retroactive cleanup: rows already classified as "Chicken Road Bonus" /
    # "Bonus Hunter" (real games, not bonuses) from before classify_bonus()
    # was fixed to require a blank `source` -- both always have their
    # provider populated in `source`, so the current classifier already
    # excludes them going forward, but already-ingested rows for recent days
    # need to be removed explicitly, or reports reading straight from
    # `bonuses` would keep showing them for weeks until they age out of the
    # 33-day window on their own. Safe to run every time (a no-op once these
    # are gone).
    cur.execute(
        "DELETE FROM bonuses WHERE bonus_name IN ('Chicken Road Bonus', 'Bonus Hunter') "
        "OR matched_category IN ('Chicken Road Bonus', 'Bonus Hunter')"
    )
    conn.commit()
    cur.execute("CREATE INDEX IF NOT EXISTS idx_bonus_user ON bonuses(user_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_bonus_name ON bonuses(bonus_name)")
    added = 0
    new_bonus_rows = []
    for f in files:
        if already_ingested(conn, f):
            print(f"  skip (already ingested): {f}")
            continue
        _, rows = load_sheet(f)
        for row in rows:
            row = clean(row)
            cur.execute(f"INSERT OR IGNORE INTO wallet_transactions VALUES ({','.join(['?']*n_cols)})", row)
            if cur.rowcount:
                added += 1
                _id, game_name, user_id = row[0], row[1], row[2]
                change_value, change_after = row[5], row[6]
                create_time, source = row[17], row[12]
                source_id = row[8]
                matched = classify_bonus(game_name, source, source_id)
                if matched:
                    new_bonus_rows.append((_id, user_id, game_name, matched, change_value, change_after, create_time, source))
        mark_ingested(conn, f)
        conn.commit()
    if new_bonus_rows:
        cur.executemany(
            "INSERT OR IGNORE INTO bonuses (id, user_id, bonus_name, matched_category, change_value, change_after, create_time, source) VALUES (?,?,?,?,?,?,?,?)",
            new_bonus_rows,
        )
        conn.commit()
    print(f"Wallet transactions: {added} rows added, {len(new_bonus_rows)} classified as bonuses")

    # Backfill: rows from files ingested in earlier runs (before classify_bonus()
    # recognized their category) never got a bonuses row, since the loop above
    # only classifies newly-inserted rows per run. Historically this re-scanned
    # the ENTIRE table every single run (16.9M+ rows and growing) -- almost all
    # of it wasted work re-checking rows already confirmed not-a-bonus under an
    # unchanged rule set. Now watermarked: only rows with id > the last
    # fully-checked id are scanned on a normal run. Whenever
    # CLASSIFY_BONUS_RULES_VERSION changes (a new/changed rule shipped), the
    # watermark is ignored for one run to fully re-scan under the new rules --
    # INSERT OR IGNORE on the shared id makes repeat scans a no-op regardless.
    cur.execute("CREATE TABLE IF NOT EXISTS backfill_state (key TEXT PRIMARY KEY, value TEXT)")
    stored_version_row = cur.execute("SELECT value FROM backfill_state WHERE key = 'rules_version'").fetchone()
    stored_version = int(stored_version_row[0]) if stored_version_row else None
    last_id_row = cur.execute("SELECT value FROM backfill_state WHERE key = 'last_backfilled_id'").fetchone()
    last_backfilled_id = int(last_id_row[0]) if last_id_row else 0

    if stored_version != CLASSIFY_BONUS_RULES_VERSION:
        print(f"Bonus classify rules changed ({stored_version} -> {CLASSIFY_BONUS_RULES_VERSION}) or first run with watermarking -- full backfill re-scan")
        backfill_rows = cur.execute(
            "SELECT id, game_name, user_id, change_value, change_after, create_time, source, source_id "
            "FROM wallet_transactions WHERE id NOT IN (SELECT id FROM bonuses)"
        ).fetchall()
    else:
        backfill_rows = cur.execute(
            "SELECT id, game_name, user_id, change_value, change_after, create_time, source, source_id "
            "FROM wallet_transactions WHERE id > ? AND id NOT IN (SELECT id FROM bonuses)",
            (last_backfilled_id,),
        ).fetchall()

    backfilled = []
    for _id, game_name, user_id, change_value, change_after, create_time, source, source_id in backfill_rows:
        matched = classify_bonus(game_name, source, source_id)
        if matched:
            backfilled.append((_id, user_id, game_name, matched, change_value, change_after, create_time, source))
    if backfilled:
        cur.executemany(
            "INSERT OR IGNORE INTO bonuses (id, user_id, bonus_name, matched_category, change_value, change_after, create_time, source) VALUES (?,?,?,?,?,?,?,?)",
            backfilled,
        )

    new_max_id = cur.execute("SELECT MAX(id) FROM wallet_transactions").fetchone()[0] or last_backfilled_id
    cur.execute(
        "INSERT OR REPLACE INTO backfill_state (key, value) VALUES ('rules_version', ?)",
        (str(CLASSIFY_BONUS_RULES_VERSION),),
    )
    cur.execute(
        "INSERT OR REPLACE INTO backfill_state (key, value) VALUES ('last_backfilled_id', ?)",
        (str(new_max_id),),
    )
    conn.commit()
    print(f"Bonus backfill: {len(backfilled)} previously-missed rows classified as bonuses (scanned {len(backfill_rows)} candidates)")
    conn.close()


def purge_old_daily_records():
    conn = sqlite3.connect(DAILY_DB)
    cur = conn.cursor()
    cutoff = f"datetime('now', '-{RETENTION_DAYS} days')"
    for table, time_col in [("deposits", "create_time"), ("withdrawals", "create_time"),
                             ("wallet_transactions", "create_time"), ("bonuses", "create_time")]:
        cur.execute(f"DELETE FROM {table} WHERE {time_col} IS NOT NULL AND datetime({time_col}) < {cutoff}")
        print(f"Purged {cur.rowcount} rows from {table} (older than {RETENTION_DAYS} days)")
    conn.commit()
    conn.execute("VACUUM")
    conn.close()


def upload_to_r2(files):
    subprocess.run([sys.executable, os.path.join(BASE, "upload_to_r2.py"), "--files"] + files, check=True)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--userlist", nargs="*", default=[])
    ap.add_argument("--deposits", nargs="*", default=[])
    ap.add_argument("--withdrawals", nargs="*", default=[])
    ap.add_argument("--wallet", nargs="*", default=[])
    ap.add_argument("--agents", nargs="*", default=[])
    ap.add_argument("--bulk-reassign", nargs="*", default=[])
    ap.add_argument("--no-upload", action="store_true")
    ap.add_argument("--no-purge", action="store_true")
    args = ap.parse_args()

    if args.userlist:
        ingest_userlist(args.userlist)
    if args.deposits:
        ingest_deposits(args.deposits)
    if args.withdrawals:
        ingest_withdrawals(args.withdrawals)
    if args.wallet:
        ingest_wallet(args.wallet)
    if args.agents:
        ingest_agents(args.agents)
    if args.bulk_reassign:
        ingest_bulk_reassign(args.bulk_reassign)

    if not args.no_purge:
        purge_old_daily_records()

    if not args.no_upload:
        # Only upload DBs that were actually touched this run -- master_userlist.db is
        # 200MB+ and rarely changes; re-uploading it on every deposits/withdrawals/wallet
        # pull wastes minutes on the scheduled pipeline for no reason.
        touched = []
        if args.userlist or args.agents or args.bulk_reassign:
            touched.append("master_userlist.db")
        if args.deposits or args.withdrawals or args.wallet:
            touched.append("daily_records.db")
        if touched:
            upload_to_r2(touched)


if __name__ == "__main__":
    main()
